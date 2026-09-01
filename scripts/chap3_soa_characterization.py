from __future__ import annotations

import json
from pathlib import Path

import matplotlib as mpl

mpl.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data" / "chap3" / "3.3_soa_characterization" / "3.3.1-2"
SCREENSHOT_DIR = ROOT / "data" / "chap3" / "3.3_soa_characterization" / "3.3.1-1"
XLSX = DATA_DIR / "不同电流下的SOA，输出光功率.xlsx"
FIGURE_BASE = ROOT / "images" / "chap3_soa_current_characterization"
SUMMARY_CSV = DATA_DIR / "processed_soa_characterization_summary.csv"
STATS_JSON = DATA_DIR / "processed_soa_characterization_stats.json"

DT = 20e-9
REPETITION_HZ = 2_000.0
PERIOD_SAMPLES = round(1.0 / (REPETITION_HZ * DT))
RESPONSIVITY_A_PER_W = 0.9
LOADED_CONVERSION_GAIN_V_PER_W = 11e3
FIXED_ATTENUATION_DB = 38.0
ATTENUATION_FACTOR = 10 ** (FIXED_ATTENUATION_DB / 10)
POWER_FACTOR_MW_PER_MV = ATTENUATION_FACTOR / LOADED_CONVERSION_GAIN_V_PER_W
EQUIVALENT_TRANSIMPEDANCE_OHM = (
    LOADED_CONVERSION_GAIN_V_PER_W / RESPONSIVITY_A_PER_W
)

SCREENSHOT_CONFIG = {
    200: ("RigolDS2.png", 0.1),
    500: ("RigolDS5.png", 0.1),
    1000: ("RigolDS10.png", 0.2),
    1500: ("RigolDS15.png", 0.5),
    2000: ("RigolDS20.png", 0.5),
}

BLUE = "#2F6B9A"
ORANGE = "#D97706"
TEAL = "#258F83"
PURPLE = "#7A5195"
GRAY = "#6B7280"
LIGHT_BLUE = "#9EC5DF"


def load_excel_data():
    ws = load_workbook(XLSX, data_only=True).active
    currents = np.array([ws.cell(r, 3).value for r in range(8, 27)], dtype=float)
    voltage_columns = (5, 7, 9, 11, 13)
    power_columns = (4, 6, 8, 10, 12)
    voltage_runs = np.array(
        [[ws.cell(r, c).value for c in voltage_columns] for r in range(8, 27)],
        dtype=float,
    )
    power_runs = np.array(
        [[ws.cell(r, c).value for c in power_columns] for r in range(8, 13)],
        dtype=float,
    )
    reverse = {
        float(ws.cell(r, 3).value): float(ws.cell(r, 5).value)
        for r in range(37, 56)
    }
    reverse_voltage = np.array([reverse[current] for current in currents])
    return currents, voltage_runs, power_runs, reverse_voltage


def csv_name(current_mA):
    if current_mA == 200:
        return "0.2.csv"
    if current_mA == 700:
        return "0.71.csv"
    return f"{current_mA / 1000:.2f}.csv"


def read_waveform(current_mA):
    path = DATA_DIR / csv_name(int(current_mA))
    return pd.read_csv(
        path,
        skiprows=1,
        header=None,
        usecols=[0],
        dtype=np.float32,
    ).iloc[:, 0].to_numpy()


def extract_pulse_metrics(signal):
    baseline = float(np.median(signal))
    cycles = signal.size // PERIOD_SAMPLES
    cycle_view = signal[: cycles * PERIOD_SAMPLES].reshape(cycles, PERIOD_SAMPLES)
    median_amplitude = float(np.median(cycle_view.max(axis=1) - baseline))
    threshold = baseline + 0.5 * median_amplitude
    indices = np.flatnonzero(signal > threshold)
    cuts = np.flatnonzero(np.diff(indices) > 1) + 1
    groups = [group for group in np.split(indices, cuts) if group.size >= 2]
    peak_indices = np.array(
        [group[np.argmax(signal[group])] for group in groups], dtype=int
    )
    pulse_peaks = signal[peak_indices] - baseline
    widths_ns = np.array([group.size * DT * 1e9 for group in groups])
    complete = (peak_indices >= 12) & (peak_indices < signal.size - 18)
    peak_indices = peak_indices[complete]
    pulse_peaks = pulse_peaks[complete]
    widths_ns = widths_ns[complete]
    intervals_us = np.diff(peak_indices) * DT * 1e6

    traces = []
    for peak_index in peak_indices:
        trace = signal[peak_index - 12 : peak_index + 19].astype(float)
        local_baseline = np.mean(np.r_[trace[:6], trace[-6:]])
        traces.append(trace - local_baseline)
    mean_trace = np.mean(np.vstack(traces), axis=0)
    time_ns = np.arange(-12, 19) * DT * 1e9
    return {
        "pulse_count": int(pulse_peaks.size),
        "peak_mean_mV": float(pulse_peaks.mean() * 1e3),
        "peak_sd_mV": float(pulse_peaks.std(ddof=1) * 1e3),
        "peak_cv_pct": float(pulse_peaks.std(ddof=1) / pulse_peaks.mean() * 100),
        "width_mean_ns": float(widths_ns.mean()),
        "width_sd_ns": float(widths_ns.std(ddof=1)),
        "period_mean_us": float(intervals_us.mean()),
        "period_sd_us": float(intervals_us.std(ddof=1)),
        "time_ns": time_ns,
        "mean_trace_V": mean_trace,
    }


def extract_screenshot_trace(current_mA):
    filename, volts_per_division = SCREENSHOT_CONFIG[current_mA]
    image = np.asarray(Image.open(SCREENSHOT_DIR / filename).convert("RGB"))
    x_pixels = np.arange(350, 920)
    y_min, y_max = 150, 650
    trace_y = np.full(x_pixels.size, np.nan)

    for index, x_pixel in enumerate(x_pixels):
        column = image[y_min:y_max, x_pixel]
        yellow = (
            (column[:, 0] > 180)
            & (column[:, 1] > 180)
            & (column[:, 2] < 120)
        )
        matches = np.flatnonzero(yellow)
        if matches.size:
            trace_y[index] = np.median(matches + y_min)

    valid = np.isfinite(trace_y)
    trace_y = np.interp(x_pixels, x_pixels[valid], trace_y[valid])
    trace_y = (
        pd.Series(trace_y)
        .rolling(window=3, center=True, min_periods=1)
        .median()
        .to_numpy()
    )

    # The screenshot grid is 50 ns/div horizontally and the vertical scale is
    # recorded per screenshot. Align the trace to the trigger at x=564 px.
    time_ns = (x_pixels - 564) * (50.0 / 111.5)
    baseline_region = (time_ns >= -180) & (time_ns <= -40)
    baseline_y = float(np.median(trace_y[baseline_region]))
    voltage = (baseline_y - trace_y) * volts_per_division / 77.0
    return time_ns, voltage


def style_axes(ax):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", color="#D1D5DB", linewidth=0.55, alpha=0.35)
    ax.tick_params(direction="out", length=3, width=0.8)


def add_panel_label(ax, label):
    ax.text(
        -0.14,
        1.07,
        label,
        transform=ax.transAxes,
        fontsize=9.5,
        fontweight="bold",
        va="top",
        ha="left",
    )


def main():
    mpl.rcParams.update(
        {
            "font.family": "sans-serif",
            "font.sans-serif": ["Arial", "DejaVu Sans"],
            "font.size": 8,
            "axes.labelsize": 8.5,
            "xtick.labelsize": 7.5,
            "ytick.labelsize": 7.5,
            "legend.fontsize": 7,
            "axes.linewidth": 0.8,
            "lines.linewidth": 1.35,
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
            "savefig.facecolor": "white",
        }
    )

    currents, voltage_runs, power_runs, reverse_voltage = load_excel_data()
    voltage_mean = voltage_runs.mean(axis=1)
    voltage_sd = voltage_runs.std(axis=1, ddof=1)
    scan_cv = voltage_sd / voltage_mean * 100
    reverse_delta_pct = (reverse_voltage - voltage_mean) / voltage_mean * 100

    converted_runs = POWER_FACTOR_MW_PER_MV * voltage_runs
    power_mean = converted_runs.mean(axis=1)
    power_sd = converted_runs.std(axis=1, ddof=1)
    direct_mean = power_runs.mean(axis=1)
    direct_sd = power_runs.std(axis=1, ddof=1)
    direct_difference_mW = power_mean[:5] - direct_mean
    direct_difference_pct = direct_difference_mW / direct_mean * 100

    raw_metrics = {}
    representative_currents = {200, 500, 1000, 1500, 2000}
    for current in currents.astype(int):
        metrics = extract_pulse_metrics(read_waveform(current))
        raw_metrics[current] = metrics
    representative_traces = {
        current: extract_screenshot_trace(current)
        for current in representative_currents
    }

    pulse_cv = np.array([raw_metrics[int(c)]["peak_cv_pct"] for c in currents])
    pulse_width = np.array([raw_metrics[int(c)]["width_mean_ns"] for c in currents])
    pulse_width_sd = np.array([raw_metrics[int(c)]["width_sd_ns"] for c in currents])

    summary = pd.DataFrame(
        {
            "soa_current_mA": currents.astype(int),
            "bpd_peak_voltage_mean_mV": voltage_mean,
            "bpd_peak_voltage_sd_mV": voltage_sd,
            "scan_cv_pct": scan_cv,
            "reverse_peak_voltage_mV": reverse_voltage,
            "reverse_delta_pct": reverse_delta_pct,
            "peak_power_mean_mW": power_mean,
            "peak_power_sd_mW": power_sd,
            "detector_input_peak_power_uW": voltage_mean
            / LOADED_CONVERSION_GAIN_V_PER_W
            * 1e3,
            "direct_meter_power_mean_mW": np.r_[direct_mean, np.full(14, np.nan)],
            "direct_meter_difference_mW": np.r_[
                direct_difference_mW, np.full(14, np.nan)
            ],
            "direct_meter_difference_pct": np.r_[
                direct_difference_pct, np.full(14, np.nan)
            ],
            "raw_complete_pulses": [raw_metrics[int(c)]["pulse_count"] for c in currents],
            "raw_peak_cv_pct": pulse_cv,
            "raw_apparent_width_mean_ns": pulse_width,
            "raw_apparent_width_sd_ns": pulse_width_sd,
            "raw_period_mean_us": [raw_metrics[int(c)]["period_mean_us"] for c in currents],
            "raw_period_sd_us": [raw_metrics[int(c)]["period_sd_us"] for c in currents],
            "power_method": ["PD-500B datasheet conversion"] * len(currents),
        }
    )
    summary.to_csv(SUMMARY_CSV, index=False, encoding="utf-8-sig")

    stats_output = {
        "power_conversion": {
            "responsivity_A_per_W": RESPONSIVITY_A_PER_W,
            "loaded_conversion_gain_V_per_W": LOADED_CONVERSION_GAIN_V_PER_W,
            "equivalent_transimpedance_ohm": EQUIVALENT_TRANSIMPEDANCE_OHM,
            "fixed_attenuation_dB": FIXED_ATTENUATION_DB,
            "attenuation_factor": ATTENUATION_FACTOR,
            "power_factor_mW_per_mV": POWER_FACTOR_MW_PER_MV,
        },
        "current_range_mA": [int(currents.min()), int(currents.max())],
        "forward_scan_repeats": int(voltage_runs.shape[1]),
        "raw_complete_pulses_per_current": [
            int(min(row["pulse_count"] for row in raw_metrics.values())),
            int(max(row["pulse_count"] for row in raw_metrics.values())),
        ],
        "voltage_mean_endpoints_mV": [float(voltage_mean[0]), float(voltage_mean[-1])],
        "power_mean_endpoints_mW": [float(power_mean[0]), float(power_mean[-1])],
        "detector_input_power_endpoints_uW": [
            float(voltage_mean[0] / LOADED_CONVERSION_GAIN_V_PER_W * 1e3),
            float(voltage_mean[-1] / LOADED_CONVERSION_GAIN_V_PER_W * 1e3),
        ],
        "direct_meter_difference_mW_range": [
            float(direct_difference_mW.min()),
            float(direct_difference_mW.max()),
        ],
        "direct_meter_difference_pct_range": [
            float(direct_difference_pct.min()),
            float(direct_difference_pct.max()),
        ],
        "scan_cv_range_pct": [float(scan_cv.min()), float(scan_cv.max())],
        "pulse_cv_range_pct": [float(pulse_cv.min()), float(pulse_cv.max())],
        "apparent_width_range_ns": [float(pulse_width.min()), float(pulse_width.max())],
        "max_abs_reverse_delta_pct": float(np.max(np.abs(reverse_delta_pct))),
    }
    STATS_JSON.write_text(
        json.dumps(stats_output, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    fig, axes = plt.subplots(2, 2, figsize=(7.25, 5.85), constrained_layout=True)
    ax_a, ax_b, ax_c, ax_d = axes.flat

    colors = mpl.colormaps["viridis"](np.linspace(0.12, 0.88, 5))
    for color, current in zip(colors, sorted(representative_currents)):
        time_ns, trace = representative_traces[current]
        ax_a.plot(
            time_ns,
            trace,
            color=color,
            linewidth=1.15,
            label=f"{current / 1000:.1f} A",
        )
    ax_a.axhline(0, color="#9CA3AF", linewidth=0.65)
    ax_a.set_xlim(-60, 160)
    ax_a.set_ylim(-0.08, 1.10)
    ax_a.set_xlabel("Time relative to trigger (ns)")
    ax_a.set_ylabel("BPD output (V)")
    ax_a.legend(frameon=False, ncol=2, loc="upper right", handlelength=1.7)
    for run in range(voltage_runs.shape[1]):
        ax_b.plot(
            currents,
            voltage_runs[:, run],
            color=LIGHT_BLUE,
            marker="o",
            markersize=2.2,
            linewidth=0.75,
            alpha=0.55,
        )
    ax_b.errorbar(
        currents,
        voltage_mean,
        yerr=voltage_sd,
        color=BLUE,
        marker="o",
        markersize=4.2,
        markeredgecolor="white",
        markeredgewidth=0.55,
        capsize=2.1,
        linewidth=1.5,
        label="Forward mean ± SD (n=5)",
        zorder=4,
    )
    ax_b.plot(
        currents,
        reverse_voltage,
        color=ORANGE,
        linestyle="--",
        marker="D",
        markerfacecolor="white",
        markeredgewidth=0.9,
        markersize=3.5,
        label="Reverse sweep",
    )
    ax_b.set_xlabel("SOA pulse current (mA)")
    ax_b.set_ylabel("BPD peak voltage (mV)")
    ax_b.set_xticks([200, 600, 1000, 1400, 1800, 2000])
    ax_b.legend(frameon=False, loc="upper left")

    jitter = np.linspace(-16, 16, power_runs.shape[1])
    for row, current in enumerate(currents[:5]):
        ax_c.scatter(
            current + jitter,
            power_runs[row],
            s=11,
            facecolor=BLUE,
            edgecolor="white",
            linewidth=0.35,
            alpha=0.58,
            zorder=2,
        )
    for row, current in enumerate(currents):
        ax_c.scatter(
            current + jitter,
            converted_runs[row],
            s=12,
            facecolor="white",
            edgecolor=ORANGE,
            linewidth=0.75,
            alpha=0.72,
            zorder=2,
        )
    ax_c.plot(currents, power_mean, color=ORANGE, linewidth=1.15, zorder=1)
    ax_c.errorbar(
        currents[:5],
        direct_mean,
        yerr=direct_sd,
        color=BLUE,
        marker="o",
        markersize=4.4,
        markeredgecolor="white",
        markeredgewidth=0.55,
        capsize=2.2,
        linewidth=1.45,
        label="JW8103A direct (n=5)",
        zorder=4,
    )
    ax_c.errorbar(
        currents,
        power_mean,
        yerr=power_sd,
        color=ORANGE,
        linestyle="--",
        marker="s",
        markerfacecolor="white",
        markeredgewidth=0.9,
        markersize=4.2,
        capsize=2.2,
        linewidth=1.4,
        label="PD-500B datasheet conversion (n=5)",
        zorder=4,
    )
    ax_c.text(
        0.97,
        0.06,
        (
            rf"$P_{{\mathrm{{SOA}}}}={POWER_FACTOR_MW_PER_MV:.4f}V_{{\mathrm{{scope}}}}$"
            + "\n"
            + rf"$G_{{50\,\Omega}}=11$ kV/W, $L={FIXED_ATTENUATION_DB:.0f}$ dB"
        ),
        transform=ax_c.transAxes,
        ha="right",
        va="bottom",
        fontsize=7,
        color="#374151",
    )
    ax_c.set_xlabel("SOA pulse current (mA)")
    ax_c.set_ylabel("Peak optical power (mW)")
    ax_c.set_xticks([200, 600, 1000, 1400, 1800, 2000])
    ax_c.legend(frameon=False, loc="upper left", ncol=1)

    line1 = ax_d.plot(
        currents,
        scan_cv,
        color=BLUE,
        marker="o",
        markersize=3.8,
        label="Scan-to-scan CV",
    )[0]
    line2 = ax_d.plot(
        currents,
        pulse_cv,
        color=ORANGE,
        linestyle="--",
        marker="s",
        markerfacecolor="white",
        markeredgewidth=0.8,
        markersize=3.8,
        label="Pulse-to-pulse CV",
    )[0]
    ax_d.set_xlabel("SOA pulse current (mA)")
    ax_d.set_ylabel("Coefficient of variation (%)")
    ax_d.set_xticks([200, 600, 1000, 1400, 1800, 2000])
    ax_d.set_ylim(0, 5.25)
    ax_d_right = ax_d.twinx()
    line3 = ax_d_right.plot(
        currents,
        pulse_width,
        color=TEAL,
        marker="^",
        markerfacecolor="white",
        markeredgewidth=0.8,
        markersize=3.8,
        linewidth=1.15,
        label="Apparent width",
    )[0]
    ax_d_right.set_ylabel("Apparent pulse width (ns)", color=TEAL)
    ax_d_right.tick_params(axis="y", colors=TEAL, direction="out", length=3)
    ax_d_right.spines["top"].set_visible(False)
    ax_d_right.spines["right"].set_color(TEAL)
    ax_d_right.set_ylim(88, 112)
    ax_d.legend(
        [line1, line2, line3],
        ["Scan-to-scan CV", "Pulse-to-pulse CV", "Apparent width"],
        frameon=False,
        loc="upper right",
    )
    ax_d_right.text(
        0.98,
        0.07,
        r"50% threshold; $\Delta t=20$ ns",
        transform=ax_d_right.transAxes,
        ha="right",
        va="bottom",
        fontsize=7.0,
        color=GRAY,
    )

    for label, ax in zip(("(a)", "(b)", "(c)", "(d)"), axes.flat):
        style_axes(ax)
        add_panel_label(ax, label)

    fig.savefig(FIGURE_BASE.with_suffix(".pdf"), bbox_inches="tight")
    fig.savefig(FIGURE_BASE.with_suffix(".png"), dpi=600, bbox_inches="tight")
    plt.close(fig)

    print(json.dumps(stats_output, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
