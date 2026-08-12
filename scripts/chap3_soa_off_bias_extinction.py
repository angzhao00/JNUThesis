from __future__ import annotations

import json
import re
from pathlib import Path

import matplotlib as mpl

mpl.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy import stats


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data" / "chap3" / "3.3_soa_characterization" / "3.3.2"
XLSX = DATA_DIR / "SOA消光比.xlsx"
PROCESSED_CSV = DATA_DIR / "processed_soa_off_bias_measurements.csv"
SUMMARY_CSV = DATA_DIR / "processed_soa_off_bias_summary.csv"
STATS_JSON = DATA_DIR / "processed_soa_off_bias_stats.json"
FIGURE_BASE = ROOT / "images" / "chap3_soa_off_bias_extinction"

BLUE = "#2F6B9A"
LIGHT_BLUE = "#9EC5DF"
ORANGE = "#D97706"
TEAL = "#258F83"
PURPLE = "#7A5195"
GRAY = "#6B7280"
LIGHT_GRAY = "#B8BDC5"


def load_measurements() -> pd.DataFrame:
    ws = load_workbook(XLSX, data_only=True).active
    blocks: list[list[dict[str, float]]] = []
    current_block: list[dict[str, float]] = []

    for row in range(9, ws.max_row + 1):
        bias_cell = ws.cell(row, 4).value
        values = [ws.cell(row, column).value for column in (5, 6, 7)]
        if bias_cell is None or any(value is None for value in values):
            if current_block:
                blocks.append(current_block)
                current_block = []
            continue

        match = re.search(r"\d+(?:\.\d+)?", str(bias_cell))
        if match is None:
            raise ValueError(f"Cannot parse off-state bias in row {row}: {bias_cell!r}")
        magnitude = float(match.group())
        current_block.append(
            {
                "bias_magnitude_V": magnitude,
                "off_bias_V": -magnitude,
                "dark_nW": float(values[0]),
                "leakage_reading_nW": float(values[1]),
                "on_peak_power_mW": float(values[2]),
            }
        )

    if current_block:
        blocks.append(current_block)
    if len(blocks) != 6 or any(len(block) != 6 for block in blocks):
        raise ValueError(
            f"Expected six complete 6-point scans, received {[len(block) for block in blocks]}"
        )

    records = []
    for scan_id, block in enumerate(blocks, start=1):
        for record in block:
            record = dict(record)
            record["scan_id"] = scan_id
            record["scan_direction"] = "reverse" if scan_id == 6 else "forward"
            records.append(record)

    data = pd.DataFrame(records)
    data["net_static_leakage_nW"] = (
        data["leakage_reading_nW"] - data["dark_nW"]
    )
    if (data["net_static_leakage_nW"] <= 0).any():
        bad = data.loc[data["net_static_leakage_nW"] <= 0]
        raise ValueError(f"Non-positive background-corrected leakage found:\n{bad}")
    data["static_extinction_ratio_dB"] = 10.0 * np.log10(
        data["on_peak_power_mW"] * 1e6 / data["net_static_leakage_nW"]
    )
    return data.sort_values(["scan_id", "bias_magnitude_V"]).reset_index(drop=True)


def summarize(data: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    fields = [
        "dark_nW",
        "leakage_reading_nW",
        "net_static_leakage_nW",
        "on_peak_power_mW",
        "static_extinction_ratio_dB",
    ]
    rows = []
    for magnitude, group in data.groupby("bias_magnitude_V", sort=True):
        forward = group[group["scan_direction"] == "forward"]
        reverse = group[group["scan_direction"] == "reverse"].iloc[0]
        row = {
            "bias_magnitude_V": float(magnitude),
            "off_bias_V": -float(magnitude),
            "all_n": int(len(group)),
            "forward_n": int(len(forward)),
        }
        for field in fields:
            all_values = group[field].to_numpy(float)
            forward_values = forward[field].to_numpy(float)
            row[f"{field}_mean"] = float(all_values.mean())
            row[f"{field}_sd"] = float(all_values.std(ddof=1))
            row[f"{field}_ci95_halfwidth"] = float(
                stats.t.ppf(0.975, len(all_values) - 1)
                * all_values.std(ddof=1)
                / np.sqrt(len(all_values))
            )
            row[f"{field}_forward_mean"] = float(forward_values.mean())
            row[f"{field}_forward_sd"] = float(forward_values.std(ddof=1))
            row[f"{field}_reverse"] = float(reverse[field])
            row[f"{field}_reverse_delta_pct"] = float(
                (reverse[field] - forward_values.mean())
                / forward_values.mean()
                * 100.0
            )
        rows.append(row)

    summary = pd.DataFrame(rows)
    min_row = summary.loc[summary["net_static_leakage_nW_mean"].idxmin()]
    max_er_row = summary.loc[summary["static_extinction_ratio_dB_mean"].idxmax()]
    zero_row = summary.loc[summary["bias_magnitude_V"] == 0].iloc[0]
    five_row = summary.loc[summary["bias_magnitude_V"] == 5].iloc[0]
    stats_output = {
        "fixed_conditions": {
            "ld_current_mA": 200,
            "tec_temperature_C": 30,
            "soa_on_pulse_current_mA": 800,
            "measurement_wavelength_nm": 1550,
            "attenuator_removed": True,
            "common_optical_reference_plane": True,
        },
        "measurement_definition": {
            "forward_scans": 5,
            "reverse_scans": 1,
            "biases_V": [0, -1, -2, -3, -4, -5],
            "net_leakage": "power-meter reading minus paired laser-off background",
            "extinction_ratio": "10*log10(on-state peak power/net static leakage)",
        },
        "dark_background_all_nW": {
            "mean": float(data["dark_nW"].mean()),
            "sd": float(data["dark_nW"].std(ddof=1)),
        },
        "minimum_net_leakage": {
            "off_bias_V": float(min_row["off_bias_V"]),
            "mean_nW": float(min_row["net_static_leakage_nW_mean"]),
            "sd_nW": float(min_row["net_static_leakage_nW_sd"]),
            "ci95_halfwidth_nW": float(
                min_row["net_static_leakage_nW_ci95_halfwidth"]
            ),
        },
        "maximum_static_extinction_ratio": {
            "off_bias_V": float(max_er_row["off_bias_V"]),
            "mean_dB": float(max_er_row["static_extinction_ratio_dB_mean"]),
            "sd_dB": float(max_er_row["static_extinction_ratio_dB_sd"]),
            "ci95_halfwidth_dB": float(
                max_er_row["static_extinction_ratio_dB_ci95_halfwidth"]
            ),
        },
        "on_peak_power_change_0_to_minus5_pct": float(
            (five_row["on_peak_power_mW_mean"] - zero_row["on_peak_power_mW_mean"])
            / zero_row["on_peak_power_mW_mean"]
            * 100.0
        ),
        "max_abs_reverse_delta_pct": {
            field: float(summary[f"{field}_reverse_delta_pct"].abs().max())
            for field in fields
            if field != "static_extinction_ratio_dB"
        },
        "max_abs_reverse_delta_static_extinction_ratio_dB": float(
            (
                summary["static_extinction_ratio_dB_reverse"]
                - summary["static_extinction_ratio_dB_forward_mean"]
            )
            .abs()
            .max()
        ),
    }
    return summary, stats_output


def style_axes(ax):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", color="#D1D5DB", linewidth=0.55, alpha=0.35)
    ax.tick_params(direction="out", length=3, width=0.8)


def add_panel_label(ax, label):
    ax.text(
        -0.14,
        1.07,
        label,
        transform=ax.transAxes,
        fontsize=9.5,
        fontweight="bold",
        va="top",
        ha="left",
    )


def add_run_points(ax, data, field, color):
    offsets = np.linspace(-0.13, 0.13, 5)
    for index, (_, run) in enumerate(
        data[data["scan_direction"] == "forward"].groupby("scan_id")
    ):
        ax.scatter(
            run["bias_magnitude_V"] + offsets[index],
            run[field],
            s=12,
            color=color,
            alpha=0.35,
            edgecolors="none",
            zorder=2,
        )


def plot_summary(data: pd.DataFrame, summary: pd.DataFrame):
    mpl.rcParams.update(
        {
            "font.family": "sans-serif",
            "font.sans-serif": ["Arial", "DejaVu Sans"],
            "font.size": 7.6,
            "axes.labelsize": 8.4,
            "xtick.labelsize": 7.2,
            "ytick.labelsize": 7.2,
            "legend.fontsize": 6.9,
            "axes.linewidth": 0.85,
            "lines.linewidth": 1.3,
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
            "savefig.facecolor": "white",
        }
    )

    fig, axes = plt.subplots(2, 2, figsize=(7.25, 5.85), constrained_layout=True)
    ax_a, ax_b, ax_c, ax_d = axes.flat
    x = summary["bias_magnitude_V"].to_numpy(float)
    reverse = data[data["scan_direction"] == "reverse"].sort_values(
        "bias_magnitude_V"
    )
    tick_labels = ["0", "−1", "−2", "−3", "−4", "−5"]

    add_run_points(ax_a, data, "leakage_reading_nW", LIGHT_BLUE)
    raw_mean = summary["leakage_reading_nW_forward_mean"].to_numpy(float)
    raw_sd = summary["leakage_reading_nW_forward_sd"].to_numpy(float)
    dark_mean = summary["dark_nW_forward_mean"].to_numpy(float)
    dark_sd = summary["dark_nW_forward_sd"].to_numpy(float)
    ax_a.errorbar(
        x,
        raw_mean,
        yerr=raw_sd,
        color=BLUE,
        marker="o",
        markersize=4.2,
        capsize=2.2,
        label="Leakage reading: forward mean ± SD",
        zorder=4,
    )
    ax_a.plot(
        x,
        reverse["leakage_reading_nW"],
        color=ORANGE,
        marker="D",
        markerfacecolor="white",
        markersize=4.0,
        linestyle="--",
        label="Leakage reading: reverse scan",
        zorder=4,
    )
    ax_a.errorbar(
        x,
        dark_mean,
        yerr=dark_sd,
        color=GRAY,
        marker="s",
        markerfacecolor="white",
        markersize=3.8,
        capsize=2.0,
        linestyle=":",
        label="Laser-off background",
        zorder=3,
    )
    ax_a.set_yscale("log")
    ax_a.set_ylabel("Power-meter reading (nW)")
    ax_a.legend(frameon=False, loc="upper right")

    add_run_points(ax_b, data, "net_static_leakage_nW", "#8BBDB7")
    net_mean = summary["net_static_leakage_nW_forward_mean"].to_numpy(float)
    net_sd = summary["net_static_leakage_nW_forward_sd"].to_numpy(float)
    ax_b.errorbar(
        x,
        net_mean,
        yerr=net_sd,
        color=TEAL,
        marker="o",
        markersize=4.2,
        capsize=2.2,
        label="Forward mean ± SD (n=5)",
        zorder=4,
    )
    ax_b.plot(
        x,
        reverse["net_static_leakage_nW"],
        color=ORANGE,
        marker="D",
        markerfacecolor="white",
        markersize=4.0,
        linestyle="--",
        label="Reverse scan",
        zorder=4,
    )
    ax_b.set_yscale("log")
    ax_b.set_ylabel("Net static leakage (nW)")
    ax_b.legend(frameon=False, loc="upper right")
    min_index = int(np.argmin(summary["net_static_leakage_nW_mean"]))
    ax_b.annotate(
        f"Minimum: {summary.iloc[min_index]['net_static_leakage_nW_mean']:.2f} nW",
        xy=(x[min_index], net_mean[min_index]),
        xytext=(2.9, 28),
        arrowprops=dict(arrowstyle="->", color=GRAY, linewidth=0.8),
        color=GRAY,
        fontsize=7.0,
    )

    add_run_points(ax_c, data, "static_extinction_ratio_dB", "#B7A1C6")
    er_mean = summary["static_extinction_ratio_dB_forward_mean"].to_numpy(float)
    er_sd = summary["static_extinction_ratio_dB_forward_sd"].to_numpy(float)
    ax_c.errorbar(
        x,
        er_mean,
        yerr=er_sd,
        color=PURPLE,
        marker="o",
        markersize=4.2,
        capsize=2.2,
        label="Forward mean ± SD (n=5)",
        zorder=4,
    )
    ax_c.plot(
        x,
        reverse["static_extinction_ratio_dB"],
        color=ORANGE,
        marker="D",
        markerfacecolor="white",
        markersize=4.0,
        linestyle="--",
        label="Reverse scan",
        zorder=4,
    )
    ax_c.set_ylabel("Static extinction ratio (dB)")
    ax_c.set_ylim(38, 86)
    ax_c.legend(frameon=False, loc="lower right")
    max_index = int(np.argmax(summary["static_extinction_ratio_dB_mean"]))
    ax_c.annotate(
        f"Maximum: {summary.iloc[max_index]['static_extinction_ratio_dB_mean']:.2f} dB",
        xy=(x[max_index], er_mean[max_index]),
        xytext=(1.8, 72),
        arrowprops=dict(arrowstyle="->", color=GRAY, linewidth=0.8),
        color=GRAY,
        fontsize=7.0,
    )

    add_run_points(ax_d, data, "on_peak_power_mW", "#A8C3D7")
    on_mean = summary["on_peak_power_mW_forward_mean"].to_numpy(float)
    on_sd = summary["on_peak_power_mW_forward_sd"].to_numpy(float)
    ax_d.errorbar(
        x,
        on_mean,
        yerr=on_sd,
        color=BLUE,
        marker="o",
        markersize=4.2,
        capsize=2.2,
        label="Forward mean ± SD (n=5)",
        zorder=4,
    )
    ax_d.plot(
        x,
        reverse["on_peak_power_mW"],
        color=ORANGE,
        marker="D",
        markerfacecolor="white",
        markersize=4.0,
        linestyle="--",
        label="Reverse scan",
        zorder=4,
    )
    ax_d.set_ylabel("On-state peak power (mW)")
    ax_d.set_ylim(199.5, 210.0)
    ax_d.legend(frameon=False, loc="upper right")
    all_zero = summary.loc[summary["bias_magnitude_V"] == 0, "on_peak_power_mW_mean"].iloc[0]
    all_five = summary.loc[summary["bias_magnitude_V"] == 5, "on_peak_power_mW_mean"].iloc[0]
    change = (all_five - all_zero) / all_zero * 100.0
    ax_d.text(
        0.04,
        0.08,
        f"0 to −5 V: {change:.2f}%",
        transform=ax_d.transAxes,
        color=GRAY,
        fontsize=7.2,
    )

    for label, ax in zip(("(a)", "(b)", "(c)", "(d)"), axes.flat):
        ax.set_xlabel("SOA off-state bias (V)")
        ax.set_xticks(x, tick_labels)
        ax.set_xlim(-0.35, 5.35)
        style_axes(ax)
        add_panel_label(ax, label)

    fig.savefig(FIGURE_BASE.with_suffix(".pdf"), bbox_inches="tight")
    fig.savefig(FIGURE_BASE.with_suffix(".png"), dpi=600, bbox_inches="tight")
    plt.close(fig)


def main():
    data = load_measurements()
    summary, stats_output = summarize(data)
    data.to_csv(PROCESSED_CSV, index=False, encoding="utf-8-sig")
    summary.to_csv(SUMMARY_CSV, index=False, encoding="utf-8-sig")
    STATS_JSON.write_text(
        json.dumps(stats_output, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    plot_summary(data, summary)
    print(json.dumps(stats_output, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
